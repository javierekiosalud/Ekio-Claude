#!/usr/bin/env python3
"""Genera un modelo de valoración en Excel con fórmulas vivas y enlazadas.

Hojas: Hipotesis, WACC, DCF, Multiplos, Marca_RFR, Clientes_MEEM, WARA, Resumen.
Todas las celdas de cálculo son fórmulas reales: al cambiar una hipótesis se
recalcula el valor. Los valores precargados son placeholders para sustituir.

Uso:
    python3 modelo_valoracion.py --salida /mnt/user-data/outputs/valoracion.xlsx \
        --nombre "Empresa S.L." --anos 5
"""

import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = "1F3864"       # cabeceras
GRIS = "F2F2F2"       # celdas de cálculo
AMARILLO = "FFF2CC"   # celdas de input (editables)
VERDE = "E2EFDA"      # resultados

H_TITULO = Font(bold=True, color="FFFFFF", size=11)
H_SECCION = Font(bold=True, size=11, color=AZUL)
F_RESULT = Font(bold=True, size=11)
BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def cabecera(ws, texto, fila=1, ancho=8):
    ws.cell(fila, 1, texto).font = H_TITULO
    for c in range(1, ancho + 1):
        ws.cell(fila, c).fill = PatternFill("solid", fgColor=AZUL)
    ws.cell(fila, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[fila].height = 22


def etiqueta(ws, fila, texto, negrita=False):
    c = ws.cell(fila, 1, texto)
    if negrita:
        c.font = H_SECCION
    return c


def inp(ws, fila, col, valor, fmt=None):
    c = ws.cell(fila, col, valor)
    c.fill = PatternFill("solid", fgColor=AMARILLO)
    c.border = BORDE
    if fmt:
        c.number_format = fmt
    return c


def calc(ws, fila, col, formula, fmt=None, resultado=False):
    c = ws.cell(fila, col, formula)
    c.fill = PatternFill("solid", fgColor=VERDE if resultado else GRIS)
    c.border = BORDE
    if resultado:
        c.font = F_RESULT
    if fmt:
        c.number_format = fmt
    return c


EUR = '#,##0'
EUR2 = '#,##0.00'
PCT = '0.00%'
X = '0.0"x"'


def hoja_hipotesis(wb, nombre, anos, ano0):
    ws = wb.create_sheet("Hipotesis")
    ws.column_dimensions["A"].width = 42
    for i in range(2, anos + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14

    cabecera(ws, f"HIPÓTESIS — {nombre}", 1, anos + 2)
    ws["A2"] = "Amarillo = input editable. Gris/verde = fórmula, no tocar."
    ws["A2"].font = Font(italic=True, size=9)

    etiqueta(ws, 4, "Ejercicio", True)
    for j in range(anos + 1):
        c = ws.cell(4, 2 + j, ano0 + j)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    ws.cell(4, 2).value = f"{ano0} (real)"

    filas = [
        (6, "Ingresos", EUR, [1_000_000] + [None] * anos),
        (7, "Crecimiento ingresos", PCT, [None] + [0.30] * anos),
        (8, "Margen bruto", PCT, [0.60] + [0.62] * anos),
        (9, "Gastos de personal / ingresos", PCT, [0.25] + [0.22] * anos),
        (10, "Marketing / ingresos", PCT, [0.15] + [0.14] * anos),
        (11, "Otros gastos operativos / ingresos", PCT, [0.10] + [0.09] * anos),
        (12, "Amortizaciones / ingresos", PCT, [0.03] + [0.03] * anos),
        (13, "CAPEX / ingresos", PCT, [0.04] + [0.04] * anos),
        (14, "NOF / ingresos", PCT, [0.12] + [0.12] * anos),
        (15, "Tipo impositivo efectivo", PCT, [0.25] * (anos + 1)),
    ]
    for fila, nombre_f, fmt, valores in filas:
        etiqueta(ws, fila, nombre_f)
        for j, v in enumerate(valores):
            if v is None:
                continue
            inp(ws, fila, 2 + j, v, fmt)
    # ingresos proyectados como fórmula
    for j in range(1, anos + 1):
        col = get_column_letter(2 + j)
        prev = get_column_letter(1 + j)
        calc(ws, 6, 2 + j, f"={prev}6*(1+{col}7)", EUR)

    etiqueta(ws, 17, "Parámetros de cierre", True)
    for fila, nombre_f, val, fmt in [
        (18, "g — crecimiento perpetuo", 0.02, PCT),
        (19, "Múltiplo de salida EV/EBITDA (contraste)", 8.0, X),
        (20, "Deuda financiera neta a fecha de valoración", 150_000, EUR),
        (21, "Activos no afectos", 0, EUR),
        (22, "Contingencias estimadas", 0, EUR),
        (23, "Descuento por falta de control (DLOC)", 0.00, PCT),
        (24, "Descuento por iliquidez (DLOM)", 0.20, PCT),
    ]:
        etiqueta(ws, fila, nombre_f)
        inp(ws, fila, 2, val, fmt)
    return ws


def hoja_wacc(wb):
    ws = wb.create_sheet("WACC")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    cabecera(ws, "COSTE DE CAPITAL", 1, 3)
    ws.cell(2, 3, "Fuente / justificación").font = Font(bold=True)

    etiqueta(ws, 3, "Coste de los recursos propios (Ke)", True)
    comps = [
        (4, "Rf — bono soberano 10a (media 12m)", 0.032, PCT, "Fuente y fecha"),
        (5, "Beta desapalancada comparables", 1.10, '0.00', "Muestra de comparables"),
        (6, "D/E objetivo", 0.25, '0.00', "Estructura normativa del sector"),
        (7, "Tipo impositivo", 0.25, PCT, ""),
    ]
    for fila, n, v, fmt, fuente in comps:
        etiqueta(ws, fila, n)
        inp(ws, fila, 2, v, fmt)
        inp(ws, fila, 3, fuente)
    etiqueta(ws, 8, "Beta reapalancada (Hamada)")
    calc(ws, 8, 2, "=B5*(1+(1-B7)*B6)", '0.00')
    for fila, n, v, fuente in [
        (9, "ERP — prima de riesgo de mercado", 0.062, "Damodaran / Fernández, indicar año"),
        (10, "SP — prima por tamaño", 0.040, "Deciles de capitalización"),
        (11, "CSRP — prima por riesgo específico", 0.030, "Desglosar por factores (ver ref. 04)"),
        (12, "CRP — riesgo país", 0.000, "Solo si aplica"),
    ]:
        etiqueta(ws, fila, n)
        inp(ws, fila, 2, v, PCT)
        inp(ws, fila, 3, fuente)
    etiqueta(ws, 13, "Ke")
    calc(ws, 13, 2, "=B4+B8*B9+B10+B11+B12", PCT, resultado=True)

    etiqueta(ws, 15, "Coste de la deuda y WACC", True)
    for fila, n, v, fmt in [(16, "Kd bruto", 0.06, PCT), (17, "Peso E/(D+E)", 0.80, PCT)]:
        etiqueta(ws, fila, n)
        inp(ws, fila, 2, v, fmt)
    etiqueta(ws, 18, "Peso D/(D+E)")
    calc(ws, 18, 2, "=1-B17", PCT)
    etiqueta(ws, 19, "Kd neto de impuestos")
    calc(ws, 19, 2, "=B16*(1-B7)", PCT)
    etiqueta(ws, 20, "WACC")
    calc(ws, 20, 2, "=B13*B17+B19*B18", PCT, resultado=True)

    etiqueta(ws, 22, "Desglose de la prima por riesgo específico (debe sumar B11)", True)
    for i, factor in enumerate([
        "Dependencia de persona clave", "Concentración de clientes",
        "Concentración de proveedores", "Dependencia de canal de captación",
        "Riesgo regulatorio", "Riesgo tecnológico",
        "Calidad de la información financiera", "Trayectoria / historial de plan vs real",
    ]):
        etiqueta(ws, 23 + i, factor)
        inp(ws, 23 + i, 2, 0.0, PCT)
        inp(ws, 23 + i, 3, "Motivo")
    etiqueta(ws, 31, "Total desglosado", True)
    calc(ws, 31, 2, "=SUM(B23:B30)", PCT)
    ws.cell(31, 3, "Debe coincidir con B11").font = Font(italic=True, size=9)
    return ws


def hoja_dcf(wb, anos, ano0):
    ws = wb.create_sheet("DCF")
    ws.column_dimensions["A"].width = 42
    for i in range(2, anos + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14
    cabecera(ws, "DESCUENTO DE FLUJOS DE CAJA (FCFF)", 1, anos + 2)

    etiqueta(ws, 3, "Ejercicio proyectado", True)
    for j in range(1, anos + 1):
        c = ws.cell(3, 1 + j, ano0 + j)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    def h(col_idx):
        return get_column_letter(2 + col_idx)  # columna en Hipotesis

    lineas = [
        (4, "Ingresos", lambda j: f"=Hipotesis!{h(j)}6", EUR),
        (5, "Margen bruto", lambda j: f"=B4*Hipotesis!{h(j)}8" if False else f"={get_column_letter(1+j)}4*Hipotesis!{h(j)}8", EUR),
        (6, "Gastos de personal", lambda j: f"=-{get_column_letter(1+j)}4*Hipotesis!{h(j)}9", EUR),
        (7, "Marketing", lambda j: f"=-{get_column_letter(1+j)}4*Hipotesis!{h(j)}10", EUR),
        (8, "Otros gastos operativos", lambda j: f"=-{get_column_letter(1+j)}4*Hipotesis!{h(j)}11", EUR),
        (9, "EBITDA", lambda j: f"=SUM({get_column_letter(1+j)}5:{get_column_letter(1+j)}8)", EUR),
        (10, "Amortizaciones", lambda j: f"=-{get_column_letter(1+j)}4*Hipotesis!{h(j)}12", EUR),
        (11, "EBIT", lambda j: f"={get_column_letter(1+j)}9+{get_column_letter(1+j)}10", EUR),
        (12, "Impuestos sobre EBIT", lambda j: f"=-{get_column_letter(1+j)}11*Hipotesis!{h(j)}15", EUR),
        (13, "NOPAT", lambda j: f"={get_column_letter(1+j)}11+{get_column_letter(1+j)}12", EUR),
        (14, "(+) Amortizaciones", lambda j: f"=-{get_column_letter(1+j)}10", EUR),
        (15, "(−) CAPEX", lambda j: f"=-{get_column_letter(1+j)}4*Hipotesis!{h(j)}13", EUR),
        (16, "NOF", lambda j: f"={get_column_letter(1+j)}4*Hipotesis!{h(j)}14", EUR),
        (17, "(−) Variación NOF",
         lambda j: f"=-({get_column_letter(1+j)}16-{get_column_letter(j)}16)" if j > 1
         else f"=-({get_column_letter(1+j)}16-Hipotesis!B6*Hipotesis!B14)", EUR),
        (18, "FCFF", lambda j: f"={get_column_letter(1+j)}13+{get_column_letter(1+j)}14+{get_column_letter(1+j)}15+{get_column_letter(1+j)}17", EUR),
        (19, "Factor de descuento", lambda j: f"=1/(1+WACC!$B$20)^({j}-0.5)", '0.0000'),
        (20, "FCFF descontado", lambda j: f"={get_column_letter(1+j)}18*{get_column_letter(1+j)}19", EUR),
    ]
    for fila, nombre_f, gen, fmt in lineas:
        etiqueta(ws, fila, nombre_f, negrita=nombre_f in ("EBITDA", "EBIT", "NOPAT", "FCFF"))
        for j in range(1, anos + 1):
            calc(ws, fila, 1 + j, gen(j), fmt)

    ult = get_column_letter(1 + anos)
    etiqueta(ws, 22, "Valor terminal", True)
    etiqueta(ws, 23, "VT por Gordon")
    calc(ws, 23, 2, f"={ult}18*(1+Hipotesis!$B$18)/(WACC!$B$20-Hipotesis!$B$18)", EUR)
    etiqueta(ws, 24, "VT por múltiplo de salida")
    calc(ws, 24, 2, f"={ult}9*Hipotesis!$B$19", EUR)
    etiqueta(ws, 25, "VT seleccionado (media)")
    calc(ws, 25, 2, "=AVERAGE(B23:B24)", EUR)
    etiqueta(ws, 26, "VT descontado")
    calc(ws, 26, 2, f"=B25*{ult}19", EUR)

    etiqueta(ws, 28, "Resultado", True)
    etiqueta(ws, 29, "VA de los flujos explícitos")
    calc(ws, 29, 2, f"=SUM(B20:{ult}20)", EUR)
    etiqueta(ws, 30, "VA del valor terminal")
    calc(ws, 30, 2, "=B26", EUR)
    etiqueta(ws, 31, "Enterprise Value")
    calc(ws, 31, 2, "=B29+B30", EUR, resultado=True)
    etiqueta(ws, 32, "Peso del VT sobre EV")
    calc(ws, 32, 2, "=B30/B31", PCT)
    ws.cell(32, 3, "Alarma si supera el 75%").font = Font(italic=True, size=9)
    etiqueta(ws, 33, "(−) Deuda financiera neta")
    calc(ws, 33, 2, "=-Hipotesis!B20", EUR)
    etiqueta(ws, 34, "(+) Activos no afectos")
    calc(ws, 34, 2, "=Hipotesis!B21", EUR)
    etiqueta(ws, 35, "(−) Contingencias")
    calc(ws, 35, 2, "=-Hipotesis!B22", EUR)
    etiqueta(ws, 36, "Equity Value (control, líquido)")
    calc(ws, 36, 2, "=B31+B33+B34+B35", EUR, resultado=True)
    etiqueta(ws, 37, "Equity Value ajustado (tras DLOC y DLOM)")
    calc(ws, 37, 2, "=B36*(1-Hipotesis!B23)*(1-Hipotesis!B24)", EUR, resultado=True)

    etiqueta(ws, 39, "Múltiplos implícitos (contraste)", True)
    etiqueta(ws, 40, "EV / EBITDA año 1")
    calc(ws, 40, 2, "=B31/B9", X)
    etiqueta(ws, 41, "EV / Ingresos año 1")
    calc(ws, 41, 2, "=B31/B4", X)
    return ws


def hoja_multiplos(wb):
    ws = wb.create_sheet("Multiplos")
    for col, w in zip("ABCDEFG", [30, 14, 14, 14, 14, 14, 30]):
        ws.column_dimensions[col].width = w
    cabecera(ws, "ENFOQUE DE MERCADO — MÚLTIPLOS COMPARABLES", 1, 7)

    heads = ["Comparable", "EV/EBITDA", "EV/Ventas", "EV/EBIT", "Crecimiento", "Margen EBITDA", "Notas"]
    for i, hd in enumerate(heads, start=1):
        c = ws.cell(3, i, hd)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=GRIS)
    for f in range(4, 12):
        for i in range(1, 8):
            inp(ws, f, i, None)

    etiqueta(ws, 13, "Mediana", True)
    for i, col in enumerate("BCDEF", start=2):
        calc(ws, 13, i, f"=IFERROR(MEDIAN({col}4:{col}11),0)", '0.00')
    etiqueta(ws, 14, "Ajuste por tamaño e iliquidez")
    inp(ws, 14, 2, -0.30, PCT)
    etiqueta(ws, 15, "Múltiplo EV/EBITDA ajustado")
    calc(ws, 15, 2, "=B13*(1+B14)", X, resultado=True)

    etiqueta(ws, 17, "Aplicación", True)
    etiqueta(ws, 18, "EBITDA normalizado de referencia")
    calc(ws, 18, 2, "=DCF!B9", EUR)
    etiqueta(ws, 19, "Enterprise Value por múltiplos")
    calc(ws, 19, 2, "=B18*B15", EUR, resultado=True)
    etiqueta(ws, 20, "Equity Value por múltiplos")
    calc(ws, 20, 2, "=B19-Hipotesis!B20+Hipotesis!B21-Hipotesis!B22", EUR, resultado=True)
    ws.cell(22, 1, "Documenta el universo inicial, los criterios de exclusión y la muestra final.").font = Font(italic=True, size=9)
    return ws


def hoja_rfr(wb, anos, ano0):
    ws = wb.create_sheet("Marca_RFR")
    ws.column_dimensions["A"].width = 42
    for i in range(2, anos + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14
    cabecera(ws, "VALORACIÓN DE MARCA — RELIEF FROM ROYALTY", 1, anos + 2)

    etiqueta(ws, 3, "Tasa de royalty", True)
    inp(ws, 3, 2, 0.04, PCT)
    etiqueta(ws, 4, "% de ingresos soportados por la marca")
    inp(ws, 4, 2, 1.00, PCT)
    etiqueta(ws, 5, "Gastos de mantenimiento de marca / ingresos")
    inp(ws, 5, 2, 0.00, PCT)
    etiqueta(ws, 6, "Prima de riesgo sobre WACC")
    inp(ws, 6, 2, 0.01, PCT)
    etiqueta(ws, 7, "Tasa de descuento de la marca")
    calc(ws, 7, 2, "=WACC!$B$20+B6", PCT)
    etiqueta(ws, 8, "Años de amortización fiscal (TAB)")
    inp(ws, 8, 2, 10, '0')

    etiqueta(ws, 10, "Ejercicio", True)
    for j in range(1, anos + 1):
        ws.cell(10, 1 + j, ano0 + j).font = Font(bold=True)
    for fila, nombre_f, gen, fmt in [
        (11, "Ingresos atribuibles", lambda j: f"=Hipotesis!{get_column_letter(2+j)}6*$B$4", EUR),
        (12, "Royalty bruto ahorrado", lambda j: f"={get_column_letter(1+j)}11*$B$3", EUR),
        (13, "(−) Gastos de mantenimiento", lambda j: f"=-{get_column_letter(1+j)}11*$B$5", EUR),
        (14, "(−) Impuestos", lambda j: f"=-({get_column_letter(1+j)}12+{get_column_letter(1+j)}13)*Hipotesis!{get_column_letter(2+j)}15", EUR),
        (15, "Flujo neto", lambda j: f"={get_column_letter(1+j)}12+{get_column_letter(1+j)}13+{get_column_letter(1+j)}14", EUR),
        (16, "Factor de descuento", lambda j: f"=1/(1+$B$7)^({j}-0.5)", '0.0000'),
        (17, "Flujo descontado", lambda j: f"={get_column_letter(1+j)}15*{get_column_letter(1+j)}16", EUR),
    ]:
        etiqueta(ws, fila, nombre_f)
        for j in range(1, anos + 1):
            calc(ws, fila, 1 + j, gen(j), fmt)

    ult = get_column_letter(1 + anos)
    etiqueta(ws, 19, "VA flujos explícitos")
    calc(ws, 19, 2, f"=SUM(B17:{ult}17)", EUR)
    etiqueta(ws, 20, "Valor terminal (marca de vida indefinida)")
    calc(ws, 20, 2, f"={ult}15*(1+Hipotesis!$B$18)/($B$7-Hipotesis!$B$18)*{ult}16", EUR)
    etiqueta(ws, 21, "Valor antes de TAB")
    calc(ws, 21, 2, "=B19+B20", EUR)
    etiqueta(ws, 22, "Factor TAB")
    calc(ws, 22, 2, "=1/(1-(Hipotesis!B15/B8)*((1-(1+B7)^-B8)/B7))", '0.0000')
    ws.cell(22, 3, "Aplicar solo si procede step-up fiscal — ver ref. 03 §8").font = Font(italic=True, size=9)
    etiqueta(ws, 23, "Valor de la marca")
    calc(ws, 23, 2, "=B21*B22", EUR, resultado=True)
    return ws


def hoja_meem(wb, anos, ano0):
    ws = wb.create_sheet("Clientes_MEEM")
    ws.column_dimensions["A"].width = 42
    for i in range(2, anos + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14
    cabecera(ws, "CARTERA DE CLIENTES — MEEM", 1, anos + 2)
    ws.cell(2, 1, "Valora SOLO la cartera existente a fecha de valoración. El crecimiento futuro es fondo de comercio.").font = Font(italic=True, size=9)

    etiqueta(ws, 3, "Ingresos año 0 de la cartera existente")
    inp(ws, 3, 2, 1_000_000, EUR)
    etiqueta(ws, 4, "Tasa de atrición anual")
    inp(ws, 4, 2, 0.25, PCT)
    etiqueta(ws, 5, "Margen EBIT atribuible")
    inp(ws, 5, 2, 0.20, PCT)
    etiqueta(ws, 6, "Prima de riesgo sobre WACC")
    inp(ws, 6, 2, 0.01, PCT)
    etiqueta(ws, 7, "Tasa de descuento de la cartera")
    calc(ws, 7, 2, "=WACC!$B$20+B6", PCT)
    etiqueta(ws, 8, "Años de amortización fiscal (TAB)")
    inp(ws, 8, 2, 10, '0')

    etiqueta(ws, 10, "CAC — cargos por activos contributivos (% s/ ingresos)", True)
    for fila, n, v in [(11, "Capital circulante", 0.005), (12, "Inmovilizado material", 0.010),
                       (13, "Plantilla ensamblada", 0.015), (14, "Marca (ver Marca_RFR)", 0.020)]:
        etiqueta(ws, fila, n)
        inp(ws, fila, 2, v, PCT)
    etiqueta(ws, 15, "Total CAC")
    calc(ws, 15, 2, "=SUM(B11:B14)", PCT)

    etiqueta(ws, 17, "Ejercicio", True)
    for j in range(1, anos + 1):
        ws.cell(17, 1 + j, ano0 + j).font = Font(bold=True)
    for fila, nombre_f, gen, fmt in [
        (18, "Supervivencia de la cartera", lambda j: f"=(1-$B$4)^{j}", PCT),
        (19, "Ingresos de la cartera existente", lambda j: f"=$B$3*{get_column_letter(1+j)}18", EUR),
        (20, "EBIT atribuible", lambda j: f"={get_column_letter(1+j)}19*$B$5", EUR),
        (21, "(−) Impuestos", lambda j: f"=-{get_column_letter(1+j)}20*Hipotesis!{get_column_letter(2+j)}15", EUR),
        (22, "NOPAT atribuible", lambda j: f"={get_column_letter(1+j)}20+{get_column_letter(1+j)}21", EUR),
        (23, "(−) Contributory asset charges", lambda j: f"=-{get_column_letter(1+j)}19*$B$15", EUR),
        (24, "Exceso de beneficio", lambda j: f"={get_column_letter(1+j)}22+{get_column_letter(1+j)}23", EUR),
        (25, "Factor de descuento", lambda j: f"=1/(1+$B$7)^({j}-0.5)", '0.0000'),
        (26, "Exceso descontado", lambda j: f"={get_column_letter(1+j)}24*{get_column_letter(1+j)}25", EUR),
    ]:
        etiqueta(ws, fila, nombre_f)
        for j in range(1, anos + 1):
            calc(ws, fila, 1 + j, gen(j), fmt)

    ult = get_column_letter(1 + anos)
    etiqueta(ws, 28, "VA del exceso de beneficio")
    calc(ws, 28, 2, f"=SUM(B26:{ult}26)", EUR)
    etiqueta(ws, 29, "Factor TAB")
    calc(ws, 29, 2, "=1/(1-(Hipotesis!B15/B8)*((1-(1+B7)^-B8)/B7))", '0.0000')
    etiqueta(ws, 30, "Valor de la cartera de clientes")
    calc(ws, 30, 2, "=B28*B29", EUR, resultado=True)
    ws.cell(32, 1, "Si la vida remanente excede el horizonte, amplía las columnas hasta agotar la cohorte.").font = Font(italic=True, size=9)
    return ws


def hoja_wara(wb):
    ws = wb.create_sheet("WARA")
    for col, w in zip("ABCDE", [38, 18, 14, 14, 40]):
        ws.column_dimensions[col].width = w
    cabecera(ws, "RECONCILIACIÓN WARA / WACC", 1, 5)
    ws.cell(2, 1, "Control de coherencia del PPA. La WARA debe aproximarse al WACC (±100 pb).").font = Font(italic=True, size=9)

    for i, hd in enumerate(["Activo", "Valor razonable", "Peso", "Tasa exigida", "Justificación de la tasa"], 1):
        c = ws.cell(4, i, hd)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=GRIS)

    activos = [
        ("Capital circulante", None, 0.03),
        ("Inmovilizado material", None, 0.06),
        ("Plantilla ensamblada", None, 0.12),
        ("Marca", "=Marca_RFR!B23", None),
        ("Cartera de clientes", "=Clientes_MEEM!B30", None),
        ("Tecnología / patentes", None, 0.18),
        ("Fondo de comercio (residuo)", None, 0.22),
    ]
    fila = 5
    for nombre_a, formula, tasa in activos:
        etiqueta(ws, fila, nombre_a)
        if formula:
            calc(ws, fila, 2, formula, EUR)
            inp(ws, fila, 4, 0.15, PCT)
        else:
            inp(ws, fila, 2, 0, EUR)
            inp(ws, fila, 4, tasa, PCT)
        calc(ws, fila, 3, f"=IFERROR(B{fila}/$B$13,0)", PCT)
        inp(ws, fila, 5, "")
        fila += 1

    etiqueta(ws, 13, "Total activos", True)
    calc(ws, 13, 2, "=SUM(B5:B11)", EUR)
    calc(ws, 13, 3, "=SUM(C5:C11)", PCT)
    etiqueta(ws, 14, "WARA")
    calc(ws, 14, 2, "=SUMPRODUCT(C5:C11,D5:D11)", PCT, resultado=True)
    etiqueta(ws, 15, "WACC")
    calc(ws, 15, 2, "=WACC!B20", PCT)
    etiqueta(ws, 16, "Diferencia (pb)")
    calc(ws, 16, 2, "=(B14-B15)*10000", '0')
    calc(ws, 16, 3, '=IF(ABS(B16)<=100,"COHERENTE","REVISAR HIPÓTESIS")')
    return ws


def hoja_resumen(wb, nombre):
    ws = wb.create_sheet("Resumen", 0)
    for col, w in zip("ABCD", [46, 18, 18, 18]):
        ws.column_dimensions[col].width = w
    cabecera(ws, f"SÍNTESIS DE VALORACIÓN — {nombre}", 1, 4)

    for i, hd in enumerate(["Método", "Mínimo", "Central", "Máximo"], 1):
        c = ws.cell(3, i, hd)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=GRIS)

    etiqueta(ws, 4, "DCF")
    calc(ws, 4, 2, "=DCF!B37*0.85", EUR)
    calc(ws, 4, 3, "=DCF!B37", EUR)
    calc(ws, 4, 4, "=DCF!B37*1.15", EUR)
    etiqueta(ws, 5, "Múltiplos comparables")
    calc(ws, 5, 2, "=Multiplos!B20*0.85", EUR)
    calc(ws, 5, 3, "=Multiplos!B20", EUR)
    calc(ws, 5, 4, "=Multiplos!B20*1.15", EUR)
    etiqueta(ws, 6, "Patrimonial ajustado (suelo)")
    for c in (2, 3, 4):
        inp(ws, 6, c, 0, EUR)
    etiqueta(ws, 7, "Suma de intangibles + tangibles (contraste)")
    calc(ws, 7, 3, "=WARA!B13", EUR)

    etiqueta(ws, 9, "Ponderación", True)
    for fila, met, peso in [(10, "Peso DCF", 0.60), (11, "Peso múltiplos", 0.25), (12, "Peso patrimonial", 0.15)]:
        etiqueta(ws, fila, met)
        inp(ws, fila, 2, peso, PCT)
    etiqueta(ws, 13, "Suma de pesos")
    calc(ws, 13, 2, "=SUM(B10:B12)", PCT)

    etiqueta(ws, 15, "CONCLUSIÓN DE VALOR", True)
    etiqueta(ws, 16, "Valor central ponderado")
    calc(ws, 16, 2, "=C4*B10+C5*B11+C6*B12", EUR, resultado=True)
    etiqueta(ws, 17, "Rango bajo")
    calc(ws, 17, 2, "=B4*B10+B5*B11+B6*B12", EUR, resultado=True)
    etiqueta(ws, 18, "Rango alto")
    calc(ws, 18, 2, "=D4*B10+D5*B11+D6*B12", EUR, resultado=True)

    etiqueta(ws, 20, "Controles de racionalidad", True)
    controles = [
        ("Peso del valor terminal sobre EV", "=DCF!B32", '=IF(DCF!B32>0.75,"REVISAR","OK")'),
        ("g perpetua", "=Hipotesis!B18", '=IF(Hipotesis!B18>0.03,"REVISAR","OK")'),
        ("g < WACC", "=WACC!B20-Hipotesis!B18", '=IF(WACC!B20<=Hipotesis!B18,"ERROR","OK")'),
        ("EV/EBITDA implícito vs comparables", "=DCF!B40", '=IF(Multiplos!B15=0,"—",IF(ABS(DCF!B40-Multiplos!B15)/Multiplos!B15>0.4,"REVISAR","OK"))'),
        ("WARA vs WACC", "=WARA!B16", '=IF(ABS(WARA!B16)<=100,"OK","REVISAR")'),
    ]
    for i, (n, f, test) in enumerate(controles):
        etiqueta(ws, 21 + i, n)
        calc(ws, 21 + i, 2, f, '0.00')
        calc(ws, 21 + i, 3, test)
    return ws


def main():
    p = argparse.ArgumentParser(description="Genera un modelo de valoración en Excel")
    p.add_argument("--salida", default="valoracion.xlsx")
    p.add_argument("--nombre", default="Sociedad valorada")
    p.add_argument("--anos", type=int, default=5, help="años de proyección explícita")
    p.add_argument("--ano-base", type=int, default=2025)
    args = p.parse_args()

    wb = Workbook()
    wb.remove(wb.active)
    hoja_hipotesis(wb, args.nombre, args.anos, args.ano_base)
    hoja_wacc(wb)
    hoja_dcf(wb, args.anos, args.ano_base)
    hoja_multiplos(wb)
    hoja_rfr(wb, args.anos, args.ano_base)
    hoja_meem(wb, args.anos, args.ano_base)
    hoja_wara(wb)
    hoja_resumen(wb, args.nombre)
    wb.save(args.salida)
    print(f"Modelo generado: {args.salida}")


if __name__ == "__main__":
    main()
